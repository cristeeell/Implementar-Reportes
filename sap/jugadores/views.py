from django.http import HttpResponse
from django.forms import modelform_factory
from django.shortcuts import render, redirect, get_object_or_404
from django.template import loader
from openpyxl.workbook import Workbook

from jugadores.forms import JugadorFormulario
from jugadores.models import Jugador
def agregar_jugador(request):
    pagina = loader.get_template('agregar_jugador.html')
    if request.method == 'GET':
        formulario = JugadorFormulario
    elif request.method == 'POST':
        formulario = JugadorFormulario(request.POST)
        if formulario.is_valid():
            formulario.save()
            return redirect('jugadores')
    datos = {'formulario':formulario}
    return HttpResponse(pagina.render(datos,request))

def modificar_jugador(request, jugador_id):
    pagina = loader.get_template('modificar_jugador.html')
    jugador = get_object_or_404(Jugador, pk=jugador_id)
    if request.method == 'GET':
        formulario = JugadorFormulario(instance=jugador)
    elif request.method == 'POST':
        formulario = JugadorFormulario(request.POST, instance=jugador)
        if formulario.is_valid():
            formulario.save()
            return redirect('jugadores')
    datos = {'formulario': formulario}
    return HttpResponse(pagina.render(datos, request))

def ver_jugador(request, jugador_id):
    # jugadores = jugadores.objects.get(pk=id)
    jugador = get_object_or_404(Jugador, pk=jugador_id)
    datos = {'jugador':jugador}
    #print(jugador)
    pagina = loader.get_template('ver_jugador.html')
    return HttpResponse(pagina.render(datos, request))

def eliminar_jugador(request, jugador_id):
    jugador = get_object_or_404(Jugador, pk=jugador_id)
    if jugador:
        jugador.delete()
        return redirect('jugadores')

# ReportePersonasExcel():
# Usamos el método get para generar el archivo excel
def generar_reporte(request, jugador=None, *args, **kwargs):
    jugadores = jugador.objects.order_by('nombre', 'sexo')
    wb = Workbook()
    ws = wb.active
    ws['B1'] = 'REPORTE DE JUGADORES'
    ws.merge_cells('B1:G1')
    ws['B3'] = 'NOMBRE'
    ws['C3'] = 'SEXO'
    ws['D3'] = 'EMAIL'
    ws['E3'] = 'JUEGO PREFERIDO'
    ws['F3'] = 'PLATAFORMA PREFERIDA'
    ws['G3'] = 'GENEROS PREFERIDOS'
    ws['H3'] = 'MODOS PREFERIDOS'
    ws['I3'] = 'FECHAS DE INICIO'
    ws['J3'] = 'ACTIVO'
    cont = 9
    for jugadores in jugadores:
        ws.cell(row=cont, column=2).value = jugadores.nombre
        ws.cell(row=cont, column=3).value = jugadores.sexo
        ws.cell(row=cont, column=4).value = jugadores.email
        ws.cell(row=cont, column=5).value = jugadores.juego_preferido
        ws.cell(row=cont, column=6).value = jugadores.plataforma_preferida
        ws.cell(row=cont, column=7).value = jugadores.generos_preferidos
        ws.cell(row=cont, column=8).value = jugadores.modos_preferidos
        ws.cell(row=cont, column=9).value = jugadores.fecha_de_inicio
        ws.cell(row=cont, column=10).value = jugadores.activo
        cont = cont + 1
# nombre del archivo
    nombre_archivo = "ReporteJugadoresExcel.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response
